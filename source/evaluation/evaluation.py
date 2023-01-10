# Code provided by Philipp Erler
# Code slightly adapted for this thesis
import argparse

import numpy as np
import os
import typing
import sys
import scipy.spatial as spatial
import pandas as pd
from pathlib import Path
from pysdf import SDF

import trimesh
import trimesh.sample

from source.util import dir_utils
from source.util import mesh_preprocess_operations

def sample_mesh(mesh_file, num_samples, rejection_radius=None):
    try:
        mesh = trimesh.load(mesh_file)
        # Todo: try with and without preprocess to see what brings better restults
        mesh = mesh_preprocess_operations.normalize_mesh(mesh)
        mesh = mesh_preprocess_operations.translate_to_origin(mesh)
    except:
        return np.zeros((0, 3))
    samples, face_indices = trimesh.sample.sample_surface_even(mesh, num_samples, rejection_radius)
    return samples

def chamfer_distance(file_in, file_ref, samples_per_model, num_processes=1):
    # http://graphics.stanford.edu/courses/cs468-17-spring/LectureSlides/L14%20-%203d%20deep%20learning%20on%20point%20cloud%20representation%20(analysis).pdf

    new_mesh_samples = sample_mesh(file_in, samples_per_model, rejection_radius=0.0)
    ref_mesh_samples = sample_mesh(file_ref, samples_per_model, rejection_radius=0.0)

    if new_mesh_samples.shape[0] == 0 or ref_mesh_samples.shape[0] == 0:
        return file_in, file_ref, -1.0

    leaf_size = 100
    sys.setrecursionlimit(int(max(1000, round(new_mesh_samples.shape[0] / leaf_size))))
    kdtree_new_mesh_samples = spatial.cKDTree(new_mesh_samples, leaf_size)
    kdtree_ref_mesh_samples = spatial.cKDTree(ref_mesh_samples, leaf_size)

    ref_new_dist, corr_new_ids = kdtree_new_mesh_samples.query(ref_mesh_samples, 1, workers=num_processes)
    new_ref_dist, corr_ref_ids = kdtree_ref_mesh_samples.query(new_mesh_samples, 1, workers=num_processes)

    ref_new_dist_sum = np.sum(ref_new_dist)
    new_ref_dist_sum = np.sum(new_ref_dist)
    chamfer_dist = ref_new_dist_sum + new_ref_dist_sum

    return file_in, file_ref, chamfer_dist

def get_signed_distance_pysdf(in_mesh: trimesh.Trimesh, query_pts_ms: np.ndarray):

    sdf = SDF(in_mesh.vertices, in_mesh.faces)
    dists_ms = sdf(query_pts_ms)

    return dists_ms

def intersection_over_union(file_in, file_ref, num_samples, num_dims=3):
    # https://learnopencv.com/intersection-over-union-iou-in-object-detection-and-segmentation/

    rng = np.random.default_rng(seed=42)
    samples = rng.random(size=(num_samples, num_dims)) * 2.0 - 1.0

    try:
        mesh_in = trimesh.load(file_in)
        mesh_ref = trimesh.load(file_ref)
    except:
        return file_in, file_ref, np.nan

    sdf_in = get_signed_distance_pysdf(mesh_in, samples)
    sdf_ref = get_signed_distance_pysdf(mesh_ref, samples)

    occ_in = sdf_in > 0.0
    occ_ref = sdf_ref > 0.0

    intersection = np.logical_and(occ_in, occ_ref)
    union = np.logical_or(occ_in, occ_ref)
    intersection_sum = np.sum(intersection)
    union_sum = np.sum(union)

    if union_sum == 0.0:
        iou = 0.0
    else:
        iou = intersection_sum / union_sum

    return file_in, file_ref, iou

# Important -> which metric chamfer distance per kd tree point in other mesh sum
# iou zufälliges mesh, bounding box 1, nicht so efficient
# result/{.ply
# num_samples ein paar seconds pro vergleich, ergebnisse um ein paar procent varianz
def get_metric_mesh(pmf, shape_list: list,
gt_mesh_files: list, num_samples: int,
                    metric: typing.Literal['chamfer', 'iou'] = 'chamfer') \
        -> typing.Iterable[np.ndarray]:
    cd_list = []
    for sni, shape_name in enumerate(shape_list):
        gt_mesh_file = gt_mesh_files[sni]
        mesh_file = os.path.join(pmf, shape_name)
        if os.path.isfile(mesh_file) and os.path.isfile(gt_mesh_file):
            if metric == 'chamfer':
                file_in, file_ref, metric_result = chamfer_distance(
                    file_in=mesh_file, file_ref=gt_mesh_file, samples_per_model=num_samples)
            elif metric == 'iou':
                file_in, file_ref, metric_result = intersection_over_union(
                    file_in=mesh_file, file_ref=gt_mesh_file, num_samples=num_samples)
            else:
                raise ValueError()
        elif not os.path.isfile(mesh_file):
            metric_result = np.nan
            # raise FileExistsError()
        elif not os.path.isfile(gt_mesh_file):
            raise FileExistsError()
        else:
            raise NotImplementedError()
        cd_list.append(metric_result)
    return cd_list

def make_excel_file_comparison(cd_pred_list, human_readable_results, output_file, result_file_templates, val_set,
                               low_metrics_better=True):
    # try https://realpython.com/openpyxl-excel-spreadsheets-python/

    filename = Path(output_file)
    dir_utils.create_general_folder(filename.parents[0])

    # check if writable
    try:
        f = open(output_file, 'w')
        f.close()
    except:
        raise OSError('File {} is already open'.format(output_file))

    # one shape per line, dataset per column
    cd_pred = np.array(cd_pred_list).transpose()
    data_headers_human_readable = ['Shape'] + [hr for hr in human_readable_results]
    data_headers = [''] + [rft for rft in result_file_templates]
    data_body = [[val_set[i]] + cd_pred[i].tolist() for i in range(len(val_set))]
    data = [data_headers_human_readable] + [data_headers] + data_body
    df = pd.DataFrame(data=data)

    # simple export
    # df.to_excel(output_file)

    # export with conditional formatting and average
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils.cell import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=True, header=False):
        ws.append(r)
    ws.freeze_panes = 'C3'

    if low_metrics_better:
        start_color = 'FF00AA00'
        end_color = 'FFAA0000'
    else:
        end_color = 'FF00AA00'
        start_color = 'FFAA0000'
    rule = ColorScaleRule(start_type='percentile', start_value=0, start_color=start_color,
                          mid_type='percentile', mid_value=50, mid_color='FFFFFFFF',
                          end_type='percentile', end_value=100, end_color=end_color)
    ws.conditional_formatting.add('C2:ZZ1000', rule)

    top_row = 2
    bottom_row = len(val_set) + 3
    for di in range(len(result_file_templates)):
        column = 3 + di
        column_letter = get_column_letter(column)
        ws.cell(row=bottom_row + 1, column=column).value = '=AVERAGE({}{}:{}{})'.format(
            column_letter, top_row, column_letter, bottom_row)
    ws.cell(row=bottom_row + 1, column=2).value = 'AVG'
    wb.save(output_file)

def make_quantitative_comparison(
        shape_names: typing.Sequence[str], gt_mesh_files: typing.Sequence[str],
        result_headers: typing.Sequence[str], result_file_templates: typing.Sequence[str],
        comp_output_dir: str, num_samples=10000):

    iou_pred_list = []
    cd_pred_list = []
    for pmf in result_file_templates:
        iou_pred_list.append(get_metric_mesh(pmf, shape_names, gt_mesh_files, num_samples, 'iou'))
        cd_pred_list.append(get_metric_mesh(pmf, shape_names, gt_mesh_files, num_samples, 'chamfer'))
    iou_output_file = os.path.join(comp_output_dir, 'iou.xlsx')
    make_excel_file_comparison(
        iou_pred_list, result_headers, iou_output_file,
        result_file_templates, shape_names, low_metrics_better=False)

    cd_output_file = os.path.join(comp_output_dir, 'chamfer_distance.xlsx')
    make_excel_file_comparison(cd_pred_list, result_headers, cd_output_file, result_file_templates, shape_names)

    return cd_pred_list, iou_pred_list

def run(input_dir, comp_dir, output_dir, result_headers):
    if not os.path.exists(input_dir) or not os.path.exists(comp_dir):
        raise Exception("Input dir {} or ground truth dir {} does not exist".format(input_dir, comp_dir))

    result_file_templates = []
    comp_files = []
    input_files = []
    for root, dirs, files in os.walk(input_dir):
        for dir in dirs:
            result_file_templates.append(os.path.join(root, dir))
    for root, dirs, files in os.walk(comp_dir):
        for file in files:
            if file.endswith('.ply'):
                comp_files.append(os.path.join(root, file))
                input_files.append(file)

    dir_utils.create_general_folder(output_dir)
    make_quantitative_comparison(input_files, comp_files, result_headers, result_file_templates, output_dir)

def diff_ars(args):
    run(args.input_dir, args.comp_dir, args.output_dir, args.result_headers)

def main(args):
    parser = argparse.ArgumentParser(prog="evaluation")
    parser.add_argument("--input_dir", type=str, default="", help="Directory that holds predicted meshes.")
    parser.add_argument("--comp_dir", type=str, default='..\\..\\resources\\topology_meshes', help="Directory that holds ground truth meshes.")
    parser.add_argument("--output_dir", type=str, default='..\\..\\output\\evaluation', help="Directory where excel output.")
    parser.add_argument("--result_headers", type=list, default=["normal", "depth+normal"], help="Headers for results in excel file")
    args = parser.parse_args(args)
    diff_ars(args)

if __name__ == '__main__':
    params = [
        '--input_dir', '..\\..\\resources\\eval\\test_evalscript\\input',
        '--comp_dir', '..\\..\\resources\\eval\\test_evalscript\\ground_truth'
        ]
    main(params)